import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io

def ajustar_nome_aba(nome):
    """Garante que o nome da aba não exceda 31 caracteres e remove caracteres inválidos."""
    if nome is None:
        raise ValueError("Nome não pode ser nulo")
    
    # Remove caracteres inválidos (como '/', ':', '?', '*', etc.)
    nome = re.sub(r'[\/:*?"<>|]', '', nome)
    
    # Trunca o nome para 31 caracteres, caso seja necessário
    return nome[:31]

def detectar_tabela(df_ref):
    """Detecta automaticamente a tabela no DataFrame baseado em critérios como a presença de dados válidos em várias colunas."""
    for i, row in df_ref.iterrows():
        if row.notna().sum() >= 2 and row.apply(lambda x: isinstance(x, (str, int, float))).all():
            return i  # Retorna o índice da linha que parece ser o cabeçalho
    return None

def preparar_aba(df_ref, tipo_amostra):
    """Processa uma aba do arquivo de referência, detectando a tabela e o cabeçalho automaticamente."""
    if df_ref is None or df_ref.empty:
        return pd.DataFrame(columns=['Parametros', 'VI', 'Unidade', 'Tipo de Amostra', 'Grupo'])

    df_ref = df_ref.reset_index(drop=True)
    cabecalho_index = detectar_tabela(df_ref)

    if cabecalho_index is not None:
        header_row = df_ref.iloc[cabecalho_index]
        if len(header_row) == len(df_ref.columns):
            df_ref.columns = header_row
        else:
            raise ValueError("Número de colunas e valores na linha de cabeçalho não coincidem.")

        df_ref = df_ref.iloc[cabecalho_index + 1:].reset_index(drop=True).dropna(axis=1, how='all')

    df_ref['Tipo de Amostra'] = tipo_amostra
    df_ref = converter_colunas_para_numeros(df_ref, ['VI'])  # Converter coluna VI para decimal
    return df_ref

def processar_referencias(ref_abas):
    """Processa todas as abas do arquivo de referência em um único DataFrame consolidado."""
    if not ref_abas or len(ref_abas) == 0:
        return pd.DataFrame(columns=['Parametros', 'VI', 'Unidade', 'Tipo de Amostra', 'Grupo'])

    referencias = []
    for aba, df_ref in ref_abas.items():
        tipo_amostra = aba  # Pode ser usado como referência do tipo
        df_preparado = preparar_aba(df_ref, tipo_amostra)
        referencias.append(df_preparado)

    return pd.concat(referencias, ignore_index=True)

def converter_colunas_para_numeros(df, colunas):
    """Converte colunas especificadas para números decimais, substituindo valores não numéricos por NaN."""
    for coluna in colunas:
        if coluna in df.columns:
            # Remover caracteres não numéricos
            df[coluna] = df[coluna].astype(str).str.replace('[^0-9.,-]', '', regex=True)
            # Substituir vírgulas por pontos para padronizar decimais
            df[coluna] = df[coluna].str.replace(',', '.')
            # Converter para números
            df[coluna] = pd.to_numeric(df[coluna], errors='coerce')
    return df

def comparar_resultados(df_base, df_ref):
    """Compara os valores da coluna 'Resultado' com os valores de referência e preenche a coluna 'Grupo' com base na tabela de referência."""
    if df_base is None or df_base.empty:
        raise ValueError("O DataFrame principal está vazio.")

    if df_ref is None or df_ref.empty:
        raise ValueError("O DataFrame de referência está vazio.")

    # Converter colunas para números
    df_base = converter_colunas_para_numeros(df_base, ['Resultado'])
    df_ref = converter_colunas_para_numeros(df_ref, ['VI'])

    # Adicionar colunas para os resultados
    df_base['Comparação'] = None
    df_base['Valor de Referência'] = None
    df_base['Unidade Divergente'] = None
    df_base['Grupo'] = None  # Adiciona a coluna Grupo

    for i, row in df_base.iterrows():
        analise = row['Análise']
        resultado = row['Resultado']
        unidade_base = row.get('Unidade', None)

        # Filtrar referências compatíveis com a análise
        ref_compatível = df_ref[df_ref['Parametros'] == analise]

        if not ref_compatível.empty:
            for _, ref_row in ref_compatível.iterrows():
                valor_ref = ref_row['VI']
                unidade_ref = ref_row.get('Unidade', None)

                # Adicionando o valor do Grupo da tabela de referência
                grupo_ref = ref_row.get('Grupo', None)

                # Verificar unidade
                unidade_divergente = None
                if unidade_base and unidade_ref and unidade_base != unidade_ref:
                    unidade_divergente = f"{unidade_base} != {unidade_ref}"

                # Comparar o valor com a referência
                if pd.notna(valor_ref):
                    if resultado < valor_ref:
                        comparacao = "Abaixo"
                    elif resultado > valor_ref:
                        comparacao = "Acima"
                    else:
                        comparacao = "Igual"

                    df_base.at[i, 'Comparação'] = comparacao
                    df_base.at[i, 'Valor de Referência'] = valor_ref
                    df_base.at[i, 'Unidade Divergente'] = unidade_divergente
                    df_base.at[i, 'Grupo'] = grupo_ref  # Atribui o valor de "Grupo" da tabela de referência
                    break
            else:
                df_base.at[i, 'Comparação'] = "Sem referência"
        else:
            df_base.at[i, 'Comparação'] = "Sem referência"
            df_base.at[i, 'Grupo'] = "Sem grupo"  # Valor padrão para Grupo

    # Remover colunas indesejadas do DataFrame final
    colunas_para_remover = [
        "Relatório de Análises", "Nº Amostra", "Proposta Comercial", 
        "Data do Recebimento", "Data da Publicação", "Previsão de Entrega", "Situação"
    ]
    df_base = df_base.drop(columns=[col for col in colunas_para_remover if col in df_base.columns], errors='ignore')

    return df_base

def salvar_em_abas(df_comparado):
    """Salva cada grupo de 'Grupo' em uma aba separada do Excel com formatação personalizada e formatação em negrito para valores de 'Resultado' com 'Comparação' igual a 'Acima'."""
    output = io.BytesIO()  # Usar BytesIO para salvar o arquivo na memória
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Agrupar por 'Grupo' e salvar cada grupo em uma aba separada
        for grupo, grupo_df in df_comparado.groupby('Grupo'):
            if pd.notna(grupo):  # Ignorar grupos nulos
                nome_aba = ajustar_nome_aba(str(grupo))  # Ajusta o nome da aba
                grupo_df.to_excel(writer, sheet_name=nome_aba, index=False)
                worksheet = writer.sheets[nome_aba]

                # Formatação do cabeçalho (primeira linha)
                for cell in worksheet["A1:K1"]:  # Ajusta o tamanho da fonte para o cabeçalho
                    for c in cell:
                        c.font = Font(bold=True, color="FFFFFF")
                        # Cor de fundo personalizada (60, 80, 130) em formato hexadecimal
                        c.fill = PatternFill(start_color="3C5082", end_color="3C5082", fill_type="solid")
                        c.alignment = Alignment(horizontal="center", vertical="center")

                # Ajustar largura das colunas automaticamente
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Obter a letra da coluna
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)  # Deixar um espaço extra
                    worksheet.column_dimensions[column].width = adjusted_width

                # Aplicar bordas em todas as células do grupo
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border

                # Aplicar negrito na coluna 'Resultado' onde 'Comparação' for 'Acima'
                for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        if cell.column_letter == 'D':  # Supondo que 'Resultado' esteja na coluna 'D'
                            comparacao_cell = worksheet.cell(row=cell.row, column=cell.column + 1)  # Coluna ao lado, "Comparação"
                            if comparacao_cell.value == "Acima":
                                cell.font = Font(bold=True)

        # Criar a aba "Demais Análises" com as linhas que têm a coluna 'Grupo' vazia
        df_demais_analises = df_comparado[df_comparado['Grupo'].isnull() | (df_comparado['Grupo'] == '')]
        if not df_demais_analises.empty:
            df_demais_analises.to_excel(writer, sheet_name="Demais Análises", index=False)
            worksheet = writer.sheets["Demais Análises"]

            # Formatação para a aba "Demais Análises"
            for cell in worksheet["A1:K1"]:
                for c in cell:
                    c.font = Font(bold=True, color="FFFFFF")
                    # Cor de fundo personalizada (60, 80, 130) em formato hexadecimal
                    c.fill = PatternFill(start_color="3C5082", end_color="3C5082", fill_type="solid")
                    c.alignment = Alignment(horizontal="center", vertical="center")

            # Ajuste da largura das colunas na aba "Demais Análises"
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Obter a letra da coluna
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Deixar um espaço extra
                worksheet.column_dimensions[column].width = adjusted_width

            # Aplicar bordas
            for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border

    output.seek(0)  # Rewind para o começo do arquivo
    return output



# Interface Streamlit
st.set_page_config(
    page_title="Conversor de Laudos",
    page_icon="https://github.com/TiagoGOliveira/labconvert/blob/main/L%20-%20Lead%20(1).png?raw=true",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Exibir logo
logo_url = "https://lead-sa.com.br/wp-content/uploads/elementor/thumbs/lead-sa-logo-qq9mfkoiu8l57gz69xhgrmnl3z2xrwh799bnd1xibk.png"
st.image(logo_url, width=200)

st.title("Conversor de Laudos Laboratoriais MyLims")
st.subheader("Carregue o arquivo exportado do MyLims e dos Valores de Referência para começar.")

uploaded_file = st.file_uploader("Arquivo Excel Resultados (export) (.xls ou .xlsx)", type=["xls", "xlsx"])
ref_file = st.file_uploader("Arquivo de referência (.xlsx)", type=["xlsx"])

if uploaded_file and ref_file:
    try:
        # Carregar arquivo principal
        df_principal = pd.read_excel(uploaded_file)
        if df_principal.empty:
            st.error("O arquivo de resultados está vazio ou não foi carregado corretamente.")
            st.stop()

        st.write("Prévia dos dados do laudo laboratorial:")
        st.dataframe(df_principal.head())

        # Carregar abas do arquivo de referência
        ref_abas = pd.read_excel(ref_file, sheet_name=None)
        if not ref_abas:
            st.error("O arquivo de referência não contém abas válidas.")
            st.stop()

        st.write("Abas do arquivo de referência:")
        st.write(list(ref_abas.keys()))

        # Processar todas as abas do arquivo de referência
        df_referencias = processar_referencias(ref_abas)
        if df_referencias.empty:
            st.error("Nenhuma referência válida encontrada.")
            st.stop()

        st.write("Prévia dos Valores de Referência:")
        st.dataframe(df_referencias.head())

        # Comparar os resultados
        df_comparado = comparar_resultados(df_principal, df_referencias)

        # Exibir uma visualização prévia dos resultados antes de gerar o download
        st.write("Resultados Tabelados Acima dos Valores de Referência:")
        df_comparado_acima = df_comparado[df_comparado['Comparação'] == 'Acima']
        st.dataframe(df_comparado_acima)

        # Salvar o arquivo de saída em memória
        output = salvar_em_abas(df_comparado)

        # Exibir o botão de download
        st.download_button(
            label="Baixar Planilha Formatada",
            data=output,
            file_name="Resultados_Projeto XXX.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Erro: {e}")
